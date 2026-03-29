from __future__ import annotations

from pathlib import Path
import sys
import tempfile
import unittest

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from packing_mvp.excel_report import write_packing_report


class ExcelReportTests(unittest.TestCase):
    def test_write_packing_report_creates_two_sheets_with_place_numbers(self) -> None:
        result_data = {
            "success": True,
            "packed_count": 2,
            "unpacked_count": 1,
            "truck": {"length_mm": 13400, "width_mm": 2350, "height_mm": 2400},
            "used_extents_mm": {"L": 3000, "W": 1200, "H": 1000},
            "placed_items": [
                {
                    "place_no": 2,
                    "item_id": "item_002",
                    "instance_id": "item_002_copy_001",
                    "name": "Second",
                    "source_kind": "manual",
                    "source_path": "",
                    "rotation": "XYZ",
                    "position_mm": {"x": 1200, "y": 0, "z": 0},
                    "dimensions_mm": {"L": 1000, "W": 800, "H": 700},
                },
                {
                    "place_no": 1,
                    "item_id": "item_001",
                    "instance_id": "item_001_copy_001",
                    "name": "First",
                    "source_kind": "step",
                    "source_path": "first.step",
                    "rotation": "XYZ",
                    "position_mm": {"x": 0, "y": 0, "z": 0},
                    "dimensions_mm": {"L": 1200, "W": 900, "H": 800},
                },
            ],
            "unplaced_items": [{"name": "Third", "quantity": 1, "source_path": "third.step"}],
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = write_packing_report(result_data, Path(tmp_dir) / "packing_report.xlsx")
            workbook = load_workbook(path)

        self.assertEqual(workbook.sheetnames, ["Упаковочный лист", "Отправочные места"])
        summary = workbook["Упаковочный лист"]
        details = workbook["Отправочные места"]
        self.assertEqual(summary["A1"].value, "Упаковочный лист")
        self.assertEqual(summary["A12"].value, 1)
        self.assertEqual(summary["B12"].value, "First")
        self.assertEqual(summary["A13"].value, 2)
        self.assertEqual(summary["A18"].value, "Third")
        self.assertEqual(summary["B18"].value, 1)
        self.assertEqual(details["A7"].value, 1)
        self.assertEqual(details["D7"].value, "First")
        self.assertEqual(details["A8"].value, 2)
        self.assertEqual(details["D8"].value, "Second")


if __name__ == "__main__":
    unittest.main()

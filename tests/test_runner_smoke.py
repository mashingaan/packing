from __future__ import annotations

import json
from pathlib import Path
import sys
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from packing_mvp.catalog import CatalogItem
from packing_mvp.runner import PackingRequest, PackingRunResult, make_default_output_dir, run_packing_job, run_packing_job_in_subprocess


def _catalog_item(path: Path, *, quantity: int = 1, dims: tuple[float, float, float] = (1000.0, 800.0, 600.0)) -> CatalogItem:
    return CatalogItem(
        item_id=f"item_{path.stem}",
        filename=path.name,
        source_path=str(path),
        detected_dims_mm=dims,
        dimensions_mm=dims,
        quantity=quantity,
    )


def _fake_render(placements, out_dir: Path, container_dims, logger=None):
    out_dir = Path(out_dir)
    (out_dir / "preview_top.png").write_bytes(b"top")
    (out_dir / "preview_side.png").write_bytes(b"side")
    return out_dir / "preview_top.png", out_dir / "preview_side.png"


def _fake_render_gif(placements, out_dir: Path, container_dims, logger=None):
    out_dir = Path(out_dir)
    gif_path = out_dir / "preview.gif"
    gif_path.write_bytes(b"GIF89a-preview")
    return gif_path


def _fake_export_scene(placements, output_step: Path, logger=None):
    output_step.write_text("packed-step", encoding="utf-8")
    return "source_models"


class _FakeQueue:
    def __init__(self) -> None:
        self.items: list[tuple[str, object]] = []

    def put(self, item: tuple[str, object]) -> None:
        self.items.append(item)


class RunnerSmokeTests(unittest.TestCase):
    def test_make_default_output_dir_is_timestamped(self) -> None:
        base = Path(r"C:\tmp\example.step")
        generated = make_default_output_dir(base)
        self.assertEqual(generated.parent, base.parent)
        self.assertTrue(generated.name.startswith("PackingResult_"))

    def test_run_packing_job_creates_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp = Path(tmp_dir)
            first = tmp / "first.step"
            second = tmp / "second.step"
            first.write_text("a", encoding="utf-8")
            second.write_text("b", encoding="utf-8")

            def fake_extract(input_path: Path, *, item_id: str, quantity: int = 1, scale: float = 1.0, logger=None):
                return _catalog_item(input_path, quantity=quantity)

            with patch("packing_mvp.runner.extract_catalog_item", side_effect=fake_extract):
                with patch("packing_mvp.runner.export_packed_scene", side_effect=_fake_export_scene):
                    with patch("packing_mvp.runner.render_previews", side_effect=_fake_render):
                        with patch("packing_mvp.runner.render_preview_gif", side_effect=_fake_render_gif):
                            result = run_packing_job(
                                PackingRequest(
                                    input_path=first,
                                    input_paths=(first, second),
                                    input_quantities=(2, 1),
                                    out_dir=tmp / "out",
                                    max_l=7000.0,
                                    max_w=2350.0,
                                    max_h=2400.0,
                                    gap=50.0,
                                ),
                                with_console=False,
                            )

            self.assertEqual(result.exit_code, 0)
            self.assertTrue((result.out_dir / "arranged.step").exists())
            self.assertTrue((result.out_dir / "preview_top.png").exists())
            self.assertTrue((result.out_dir / "preview_side.png").exists())
            self.assertTrue((result.out_dir / "preview.gif").exists())
            self.assertTrue((result.out_dir / "packing_report.xlsx").exists())
            data = json.loads(result.result_path.read_text(encoding="utf-8"))
            self.assertTrue(data["success"])
            self.assertEqual(data["packed_count"], 3)
            self.assertEqual(len(data["catalog"]), 2)
            self.assertEqual([item["place_no"] for item in data["placed_items"]], [1, 2, 3])
            workbook = load_workbook(result.out_dir / "packing_report.xlsx")
            self.assertEqual(workbook.sheetnames, ["Упаковочный лист", "Ручное заполнение", "Отправочные места"])

    def test_run_packing_job_reports_unplaced_items(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp = Path(tmp_dir)
            first = tmp / "first.step"
            first.write_text("a", encoding="utf-8")

            def fake_extract(input_path: Path, *, item_id: str, quantity: int = 1, scale: float = 1.0, logger=None):
                return _catalog_item(input_path, quantity=quantity, dims=(3000.0, 1000.0, 1000.0))

            with patch("packing_mvp.runner.extract_catalog_item", side_effect=fake_extract):
                with patch("packing_mvp.runner.export_packed_scene", side_effect=_fake_export_scene):
                    with patch("packing_mvp.runner.render_previews", side_effect=_fake_render):
                        with patch("packing_mvp.runner.render_preview_gif", side_effect=_fake_render_gif):
                            result = run_packing_job(
                                PackingRequest(
                                    input_path=first,
                                    input_paths=(first,),
                                    input_quantities=(3,),
                                    out_dir=tmp / "out",
                                    max_l=5000.0,
                                    max_w=1200.0,
                                    max_h=1200.0,
                                    gap=50.0,
                                ),
                                with_console=False,
                            )

            self.assertEqual(result.exit_code, 2)
            data = json.loads(result.result_path.read_text(encoding="utf-8"))
            self.assertFalse(data["success"])
            self.assertEqual(data["packed_count"], 1)
            self.assertEqual(data["unpacked_count"], 2)
            self.assertEqual(data["unplaced_items"][0]["quantity"], 2)
            self.assertTrue(result.placements_path.exists())
            self.assertTrue((result.out_dir / "packing_report.xlsx").exists())

    def test_run_packing_job_in_subprocess_emits_status_and_result(self) -> None:
        request = PackingRequest(input_path=Path("dummy.step"), out_dir=Path("out"))
        result = PackingRunResult(
            exit_code=0,
            out_dir=Path("out"),
            result_path=Path("out/result.json"),
            placements_path=Path("out/placements.csv"),
            log_path=Path("out/packing.log"),
            preview_top_path=None,
            preview_side_path=None,
            preview_gif_path=None,
            result_data={"status": "ok", "success": True},
        )
        events = _FakeQueue()

        def fake_run(_request, **kwargs):
            kwargs["status_callback"]("reading")
            kwargs["status_callback"]("packing")
            return result

        with patch("packing_mvp.runner.run_packing_job", side_effect=fake_run):
            run_packing_job_in_subprocess(request, events)

        self.assertEqual(events.items, [("status", "reading"), ("status", "packing"), ("done", result)])


if __name__ == "__main__":
    unittest.main()
